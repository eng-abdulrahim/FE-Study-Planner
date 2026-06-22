import openpyxl, json, re, os

PATH = r"_work/wb_copy.xlsx"
wv = openpyxl.load_workbook(PATH, data_only=True)
ws = wv["Topic Planner"]

def slug(s):
    s = re.sub(r"[^a-zA-Z0-9]+", "-", str(s)).strip("-").lower()
    return s

topics = []
for r in range(2, 87):
    name = ws.cell(r, 3).value
    if not name:
        continue
    section = ws.cell(r, 2).value
    m = re.match(r"\s*(\d+)\.(\d+)", str(name))
    sec_num = int(m.group(1)) if m else None
    code = f"{m.group(1)}.{m.group(2)}" if m else str(r)
    topics.append({
        "id": code,                              # e.g. "6.6"
        "rowOrder": r - 1,                        # original workbook order (1-based)
        "section": section,
        "sectionNumber": sec_num,
        "topicName": str(name),
        "qRange": ws.cell(r, 4).value,
        "examWeight": ws.cell(r, 5).value,
        "tier": ws.cell(r, 6).value,
        "defaultConfidence": ws.cell(r, 7).value,
        "defaultDifficulty": ws.cell(r, 8).value,
        "defaultBoredom": ws.cell(r, 9).value,
        "defaultQuickWin": ws.cell(r, 10).value,
        "recommendedDepth": ws.cell(r, 12).value,
        "plannedHours": ws.cell(r, 13).value,
        "includeDefault": ws.cell(r, 1).value,
        "defaultStatus": ws.cell(r, 16).value,
        "notes": ws.cell(r, 18).value,
    })

os.makedirs("app/src/data", exist_ok=True)

# Emit TypeScript
header = '''// AUTO-EXTRACTED from Latifah_FE_Auto_Generated_Study_Planner.xlsx (Topic Planner sheet, rows 2-86).
// Values are taken verbatim from the workbook. Do not hand-edit; re-run extraction to refresh.
import type { TopicSeed } from "../types/planner";

export const TOPICS: TopicSeed[] = '''

# clean None -> null, ensure numbers
def js(v):
    return v

with open("app/src/data/topics.ts", "w", encoding="utf-8") as f:
    f.write(header)
    f.write(json.dumps(topics, ensure_ascii=False, indent=2))
    f.write(";\n\nexport const TOTAL_TOPICS = TOPICS.length;\n")

# Also a plain JSON for reference
with open("_work/topics.json", "w", encoding="utf-8") as f:
    json.dump(topics, f, ensure_ascii=False, indent=2)

print("Extracted", len(topics), "topics -> app/src/data/topics.ts")
print("Sections:", sorted(set(t["section"] for t in topics)))
print("Tiers:", sorted(set(t["tier"] for t in topics)))
print("Total planned hours:", round(sum(t["plannedHours"] or 0 for t in topics), 1))
print("Sample:", json.dumps(topics[25], ensure_ascii=False))
