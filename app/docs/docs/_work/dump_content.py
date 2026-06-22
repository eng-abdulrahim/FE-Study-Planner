import openpyxl
from openpyxl.utils import get_column_letter

PATH = r"_work/wb_copy.xlsx"
wf = openpyxl.load_workbook(PATH, data_only=False)
wv = openpyxl.load_workbook(PATH, data_only=True)

def dump_sheet(name, max_r=None, max_c=None, only_formulas=False, label=""):
    ws = wf[name]; wsv = wv[name]
    mr = max_r or ws.max_row; mc = max_c or ws.max_column
    print(f"\n################ {name} {label}  (showing {mr}x{mc}) ################")
    for r in range(1, mr+1):
        for c in range(1, mc+1):
            cell = ws.cell(row=r, column=c)
            f = cell.value
            v = wsv.cell(row=r, column=c).value
            if f is None and v is None:
                continue
            coord = f"{get_column_letter(c)}{r}"
            if isinstance(f, str) and f.startswith("="):
                print(f"  {coord}: FORMULA {f}   => {repr(v)}")
            elif f.__class__.__name__ == "ArrayFormula":
                print(f"  {coord}: ARRAY {getattr(f,'text',f)}   => {repr(v)}")
            else:
                if not only_formulas:
                    print(f"  {coord}: {repr(f)}")

# Full small sheets
dump_sheet("_Lists")
dump_sheet("_Calc")
dump_sheet("_PlanEngine")
dump_sheet("Weekly Plan")
dump_sheet("Dashboard")

# Topic Planner: header + first 3 data rows full, then column-pattern (row2)
print("\n################ Topic Planner HEADERS (row1) ################")
ws = wf["Topic Planner"]; wsv = wv["Topic Planner"]
for c in range(1, ws.max_column+1):
    print(f"  {get_column_letter(c)}1: {repr(ws.cell(1,c).value)}")

print("\n################ Topic Planner ROW 2 (formula pattern per column) ################")
for c in range(1, ws.max_column+1):
    cell = ws.cell(2,c); v = wsv.cell(2,c).value
    print(f"  {get_column_letter(c)}2: {repr(cell.value)}   => {repr(v)}")

print("\n################ Topic Planner ROW 3 (values) ################")
for c in range(1, ws.max_column+1):
    cell = ws.cell(3,c); v = wsv.cell(3,c).value
    print(f"  {get_column_letter(c)}3: {repr(cell.value)}   => {repr(v)}")

# Count non-empty topic names + sample some names
print("\n################ Topic Planner: sample topic names (col C) + sections (col B) ################")
names = []
for r in range(2, 87):
    nm = wsv.cell(r,3).value
    sec = wsv.cell(r,2).value
    if nm:
        names.append((r, sec, nm))
print(f"  Non-empty topic rows: {len(names)} (of 85 possible)")
for r, sec, nm in names[:20]:
    print(f"   row{r}: [{sec}] {nm}")
print("   ...")
for r, sec, nm in names[-5:]:
    print(f"   row{r}: [{sec}] {nm}")
