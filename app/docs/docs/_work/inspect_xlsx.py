import json, re
import openpyxl
from openpyxl.utils import get_column_letter

PATH = r"_work/wb_copy.xlsx"

wb_f = openpyxl.load_workbook(PATH, data_only=False)   # formulas
wb_v = openpyxl.load_workbook(PATH, data_only=True)    # cached values

VOLATILE = ["TODAY", "NOW", "INDIRECT", "OFFSET", "RAND", "RANDBETWEEN", "CELL", "INFO", "AREAS"]
ERRORS = ["#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!"]

report = {}

# Workbook level
report["properties"] = {
    "creator": wb_f.properties.creator,
    "lastModifiedBy": wb_f.properties.lastModifiedBy,
    "created": str(wb_f.properties.created),
    "modified": str(wb_f.properties.modified),
    "title": wb_f.properties.title,
}
report["security"] = {
    "lockStructure": getattr(wb_f.security, "lockStructure", None) if wb_f.security else None,
    "lockWindows": getattr(wb_f.security, "lockWindows", None) if wb_f.security else None,
    "workbookPassword": getattr(wb_f.security, "workbookPassword", None) if wb_f.security else None,
}
report["calc_properties"] = {
    "calcMode": wb_f.calculation.calcMode if wb_f.calculation else None,
    "fullCalcOnLoad": wb_f.calculation.fullCalcOnLoad if wb_f.calculation else None,
}

# Defined names
dn = {}
for name, defn in wb_f.defined_names.items():
    dn[name] = {"value": defn.value, "hidden": defn.hidden, "localSheetId": defn.localSheetId}
report["defined_names"] = dn
report["defined_names_count"] = len(dn)

sheets = []
all_formulas = []  # (sheet, coord, formula)

for ws in wb_f.worksheets:
    wv = wb_v[ws.title]
    s = {}
    s["title"] = ws.title
    s["state"] = ws.sheet_state
    s["dimensions"] = ws.dimensions
    s["max_row"] = ws.max_row
    s["max_col"] = ws.max_column
    s["freeze_panes"] = ws.freeze_panes
    # protection
    prot = ws.protection
    s["protection"] = {
        "sheet": prot.sheet,
        "password_hash": prot.password if prot.sheet else None,
        "selectLockedCells": prot.selectLockedCells,
        "selectUnlockedCells": prot.selectUnlockedCells,
        "formatCells": prot.formatCells,
        "insertRows": prot.insertRows,
        "deleteRows": prot.deleteRows,
    }
    # merged cells
    s["merged_cells"] = [str(m) for m in ws.merged_cells.ranges]
    # hidden columns / rows
    hidden_cols = [c for c, d in ws.column_dimensions.items() if d.hidden]
    hidden_rows = [r for r, d in ws.row_dimensions.items() if d.hidden]
    s["hidden_columns"] = hidden_cols
    s["hidden_rows"] = hidden_rows
    col_widths = {c: d.width for c, d in ws.column_dimensions.items() if d.width}
    s["column_widths"] = col_widths

    # data validations
    dvs = []
    for dv in ws.data_validations.dataValidation:
        dvs.append({
            "type": dv.type,
            "operator": dv.operator,
            "formula1": dv.formula1,
            "formula2": dv.formula2,
            "allow_blank": dv.allow_blank,
            "showErrorMessage": dv.showErrorMessage,
            "errorTitle": dv.errorTitle,
            "error": dv.error,
            "showInputMessage": dv.showInputMessage,
            "promptTitle": dv.promptTitle,
            "prompt": dv.prompt,
            "sqref": str(dv.sqref),
        })
    s["data_validations"] = dvs

    # conditional formatting
    cfs = []
    for rng in ws.conditional_formatting:
        for rule in ws.conditional_formatting[rng]:
            cfs.append({
                "range": str(rng.sqref) if hasattr(rng, "sqref") else str(rng),
                "type": rule.type,
                "operator": rule.operator,
                "formula": list(rule.formula) if rule.formula else [],
                "priority": rule.priority,
                "dxfId": rule.dxfId,
                "text": getattr(rule, "text", None),
                "rank": getattr(rule, "rank", None),
                "stdDev": getattr(rule, "stdDev", None),
                "colorScale": str(rule.colorScale) if getattr(rule, "colorScale", None) else None,
                "dataBar": "yes" if getattr(rule, "dataBar", None) else None,
                "iconSet": str(rule.iconSet.iconSet) if getattr(rule, "iconSet", None) else None,
            })
    s["conditional_formatting"] = cfs

    # formulas + errors + cached values
    sheet_formulas = []
    error_cells = []
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                sheet_formulas.append({"coord": cell.coordinate, "formula": v})
                all_formulas.append((ws.title, cell.coordinate, v))
            # array formula objects
            elif v.__class__.__name__ == "ArrayFormula":
                txt = getattr(v, "text", str(v))
                sheet_formulas.append({"coord": cell.coordinate, "formula": "{ARRAY}" + str(txt)})
                all_formulas.append((ws.title, cell.coordinate, str(txt)))
            # cached error values
            cv = wv[cell.coordinate].value
            if isinstance(cv, str) and cv in ERRORS:
                error_cells.append({"coord": cell.coordinate, "error": cv})
    s["formula_count"] = len(sheet_formulas)
    s["formulas"] = sheet_formulas
    s["error_cells"] = error_cells
    sheets.append(s)

report["sheets"] = sheets

# Volatile + error scan across all formulas
vol_hits = {v: [] for v in VOLATILE}
for sh, coord, f in all_formulas:
    fu = f.upper()
    for v in VOLATILE:
        if re.search(r"\b" + v + r"\s*\(", fu):
            vol_hits[v].append(f"{sh}!{coord}")
report["volatile_usage"] = {k: v for k, v in vol_hits.items() if v}
report["total_formulas"] = len(all_formulas)

with open(r"_work/xlsx_report.json", "w", encoding="utf-8") as fh:
    json.dump(report, fh, ensure_ascii=False, indent=2)

# ---- Readable summary ----
print("WORKBOOK PROPERTIES:", json.dumps(report["properties"], ensure_ascii=False))
print("SECURITY:", json.dumps(report["security"], ensure_ascii=False))
print("CALC:", json.dumps(report["calc_properties"], ensure_ascii=False))
print("DEFINED NAMES COUNT:", report["defined_names_count"])
print("TOTAL FORMULAS:", report["total_formulas"])
print("VOLATILE USAGE COUNT:", {k: len(v) for k, v in report["volatile_usage"].items()})
print()
for s in sheets:
    print(f"=== SHEET: {s['title']}  [{s['state']}]  dims={s['dimensions']}  rows={s['max_row']} cols={s['max_col']}  formulas={s['formula_count']}")
    print(f"    freeze={s['freeze_panes']}  protection.sheet={s['protection']['sheet']}")
    print(f"    merged={len(s['merged_cells'])}  hidden_cols={s['hidden_columns']}  DV={len(s['data_validations'])}  CF={len(s['conditional_formatting'])}  errs={len(s['error_cells'])}")
    if s['error_cells']:
        print("    ERROR CELLS:", s['error_cells'])
print("\nDONE. JSON saved to _work/xlsx_report.json")
