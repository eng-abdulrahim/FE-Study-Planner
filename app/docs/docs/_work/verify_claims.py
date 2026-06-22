import openpyxl
from openpyxl.utils import get_column_letter
PATH = r"_work/wb_copy.xlsx"
wv = openpyxl.load_workbook(PATH, data_only=True)
wf = openpyxl.load_workbook(PATH, data_only=False)

ws = wv["Topic Planner"]
print("FULL TOPIC TABLE: row | Include | Section | Topic | Wt | Tier | Conf | Diff | Bore | QW | Suggested Hrs | Status")
tier_by_section = {}
for r in range(2, 87):
    inc = ws.cell(r,1).value; sec = ws.cell(r,2).value; nm = ws.cell(r,3).value
    wt = ws.cell(r,5).value; tier = ws.cell(r,6).value; conf = ws.cell(r,7).value
    diff = ws.cell(r,8).value; bore = ws.cell(r,9).value; qw = ws.cell(r,10).value
    sh = ws.cell(r,13).value; st = ws.cell(r,16).value
    print(f"  {r:>2} | {inc} | {sec} | {nm} | wt={wt} | T{tier} | C{conf} D{diff} B{bore} Q{qw} | {sh}h | {st}")
    tier_by_section.setdefault(sec, set()).add(tier)

print("\nTIER BY SECTION:")
for sec, tiers in tier_by_section.items():
    print(f"  {sec}: tiers={sorted(t for t in tiers if t is not None)}")

print("\nCLAIM CHECKS:")
def find_topic(sub):
    hits=[]
    for r in range(2,87):
        nm=ws.cell(r,3).value
        if nm and sub.lower() in str(nm).lower():
            hits.append((r, ws.cell(r,2).value, nm, ws.cell(r,6).value))
    return hits
for q in ["Electromagnetic","Laplace","Digital Filter","Z-Transform","Impedance","Communication Theory","Sampling"]:
    print(f"  '{q}': {find_topic(q)}")

# Fill colors for input vs output cells (yellow/green convention)
print("\nFILL COLORS (input vs output):")
wsf = wf["Dashboard"]
for coord, role in [("C20","input name"),("C21","input examdate"),("C22","input startdate"),
                    ("G20","input min"),("H20","input energy"),("I20","input mode"),
                    ("A5","output examdate"),("D5","output daysrem"),("J5","output prog"),
                    ("E11","output todaytopic")]:
    c = wsf[coord]
    fg = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
    pat = c.fill.patternType if c.fill else None
    print(f"  Dashboard!{coord} ({role}): pattern={pat} fill={fg}")
wtp = wf["Topic Planner"]
for coord, role in [("A2","input include"),("G2","input conf"),("I2","input boredom"),
                    ("N2","input completed"),("P2","input status"),("R2","input notes"),
                    ("K2","output priority"),("O2","output progress"),("Q2","output action"),
                    ("H2","Difficulty?"),("J2","QuickWin?"),("E2","AvgWt?"),("F2","Tier?"),("M2","SuggHrs?")]:
    c = wtp[coord]
    fg = c.fill.fgColor.rgb if c.fill and c.fill.fgColor else None
    pat = c.fill.patternType if c.fill else None
    print(f"  TopicPlanner!{coord} ({role}): pattern={pat} fill={fg}")

# Check ReadinessPct non-monotonic bug numerically (simulate)
print("\nREADINESS METRIC NOTE: B28 formula =", wf["_Calc"]["B28"].value)
print("PROGRESS METRIC: B27 formula =", wf["_Calc"]["B27"].value)
