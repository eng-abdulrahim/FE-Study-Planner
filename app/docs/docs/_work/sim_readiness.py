import openpyxl
PATH = r"_work/wb_copy.xlsx"
wv = openpyxl.load_workbook(PATH, data_only=True)
ws = wv["Topic Planner"]

topics = []
for r in range(2, 87):
    inc = ws.cell(r,1).value
    E = ws.cell(r,5).value or 0   # weight
    T = ws.cell(r,6).value or 3   # tier
    G = ws.cell(r,7).value or 3   # confidence
    I = ws.cell(r,9).value or 0   # boredom
    J = ws.cell(r,10).value or 0  # quick win
    M = ws.cell(r,13).value or 0  # suggested hours
    if inc != "Yes":
        continue
    tb = 5 if T==1 else (3 if T==2 else 1)
    base = round((E*0.35 + (6-G)*0.25 + J*0.2 + tb*0.15 - I*0.05), 2)
    topics.append({"row":r, "base":base, "hours":M})

# sort by base priority desc (engine studies highest first)
topics.sort(key=lambda x: -x["base"])
total_hours = sum(t["hours"] for t in topics)

def readiness(completed_set):
    num = 0.0; den = 0.0
    for i, t in enumerate(topics):
        done = i in completed_set
        K = t["base"]*0.1 if done else t["base"]
        comp = t["hours"] if done else 0.0
        num += K*comp
        den += K*t["hours"]
    return (num/den) if den else 0.0

print(f"Included topics: {len(topics)}, total suggested hours: {total_hours}")
print("\nSimulating completing topics one-by-one in priority order:")
print(f"{'#done':>6} {'%hours_done':>12} {'Readiness%':>11}")
done = set()
points = [0, 10, 20, 30, 42, 50, 60, 70, 80, 84, 85]
for k in range(len(topics)+1):
    if k in points:
        comp_hours = sum(topics[i]["hours"] for i in done)
        print(f"{k:>6} {100*comp_hours/total_hours:>11.1f}% {100*readiness(done):>10.2f}%")
    if k < len(topics):
        done.add(k)

# Demonstrate the drop at the 99%->100% transition for the single top topic
print("\nNon-monotonic demonstration (complete ONLY the #1 priority topic):")
t0 = topics[0]
# 99% done (not full -> full priority weight)
num = t0["base"]*(0.99*t0["hours"]); den_full = sum(t["base"]*t["hours"] for t in topics)
# other topics 0 completed
r99 = num/den_full
# 100% done -> weight *0.1 for this topic
den_after = t0["base"]*0.1*t0["hours"] + sum(t["base"]*t["hours"] for t in topics[1:])
num_after = t0["base"]*0.1*t0["hours"]
r100 = num_after/den_after
print(f"  Top topic base={t0['base']} hours={t0['hours']}")
print(f"  Readiness at 99% of that topic done (rest 0): {100*r99:.3f}%")
print(f"  Readiness at 100% of that topic done (rest 0): {100*r100:.3f}%")
print(f"  => Finishing the topic changes readiness by {100*(r100-r99):+.3f} percentage points")
