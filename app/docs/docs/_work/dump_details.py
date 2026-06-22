import openpyxl, json, re
from openpyxl.utils import get_column_letter

PATH = r"_work/wb_copy.xlsx"
wf = openpyxl.load_workbook(PATH, data_only=False)
wv = openpyxl.load_workbook(PATH, data_only=True)

print("################ Dashboard rows 1-19 (top KPIs) ################")
ws = wf["Dashboard"]; wsv = wv["Dashboard"]
for r in range(1, 19):
    for c in range(1, 19):
        f = ws.cell(r,c).value; v = wsv.cell(r,c).value
        if f is None and v is None: continue
        coord = f"{get_column_letter(c)}{r}"
        if isinstance(f,str) and f.startswith("="):
            print(f"  {coord}: = {f}   => {repr(v)}")
        else:
            print(f"  {coord}: {repr(f)}")

print("\n################ Weekly Plan FULL ################")
ws = wf["Weekly Plan"]; wsv = wv["Weekly Plan"]
for r in range(1, ws.max_row+1):
    for c in range(1, ws.max_column+1):
        f = ws.cell(r,c).value; v = wsv.cell(r,c).value
        if f is None and v is None: continue
        coord = f"{get_column_letter(c)}{r}"
        if isinstance(f,str) and f.startswith("="):
            print(f"  {coord}: = {f}   => {repr(v)}")
        else:
            print(f"  {coord}: {repr(f)}")

# ---- Reference analysis ----
print("\n################ REFERENCE SCAN ################")
rep = json.load(open(r"_work/xlsx_report.json", encoding="utf-8"))
all_formulas = []
for s in rep["sheets"]:
    for fo in s["formulas"]:
        all_formulas.append((s["title"], fo["coord"], fo["formula"]))

def scan(term):
    hits = [f"{sh}!{co}" for sh,co,fm in all_formulas if term in fm]
    return hits

# Dashboard input cells of interest
for label, ref in [("Exam Date C21","C21"),("Start Date C22","C22"),("Study Style C23","C23"),
                    ("Min Wkly C24","C24"),("Max Wkly C25","C25"),("Student Name C20","C20")]:
    # search for Dashboard!Cxx or bare Cxx in Dashboard formulas
    hits = [f"{sh}!{co}" for sh,co,fm in all_formulas if re.search(r"(Dashboard!\$?"+ref[0]+r"\$?"+ref[1:]+r")|(?<![A-Z0-9$])"+ref+r"(?![0-9])", fm)]
    print(f"  {label}: referenced in -> {hits if hits else 'NONE'}")

# Named range usage
print("\n  Named-range usage (defined but check if used in any formula):")
for nm in ["TiredKey","DeferKey","ReadinessPct","ProgressPct","OverallDone","OverallRem","WeeklyAvail","WeeklyPlanned","HighPriRem","RemainingHrs","TopicDepth","TopicAction","StatusList"]:
    hits = scan(nm)
    print(f"    {nm}: {len(hits)} uses -> {hits[:6]}{'...' if len(hits)>6 else ''}")

# Search any reference to 'Exam' text anywhere (cells), and countdown
print("\n  Search formulas mentioning exam/countdown/deadline:")
for sh,co,fm in all_formulas:
    if re.search(r"(?i)exam|deadline|countdown|days?left|remaining day", fm):
        print(f"    {sh}!{co}: {fm}")

print("\n################ DATA VALIDATIONS ################")
for s in rep["sheets"]:
    if s["data_validations"]:
        print(f"  -- {s['title']} --")
        for dv in s["data_validations"]:
            print(f"    sqref={dv['sqref']} type={dv['type']} op={dv['operator']} f1={dv['formula1']} f2={dv['formula2']} blank={dv['allow_blank']}")
            if dv['prompt'] or dv['error']:
                print(f"        prompt={dv['promptTitle']!r}/{dv['prompt']!r}  error={dv['errorTitle']!r}/{dv['error']!r}")
